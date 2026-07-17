import argparse
import csv
import hashlib
import json
from pathlib import Path

import cv2
import numpy as np

from qpd_hwk_simulator import (
    HwkBank,
    QpdHwkSimulator,
    quad_rggb_planes_to_mosaic,
    quad_rggb_sample_to_planes,
)

DEFAULT_ISP_PARAMS = {
    "cfa_pattern": "RGGB",
    "black_level": 64,
    "white_level": 1023,
    "bit_depth": 10,
    "iso": 100,
    "wb_gains": [2.0, 1.0, 1.6],
    "ccm_srgb_from_cam": [
        [1.0, 0.0, 0.0],
        [0.0, 1.0, 0.0],
        [0.0, 0.0, 1.0],
    ],
}

QPD_OUTPUT_LEVELS = {
    "bit_depth": 10,
    "black_level": 64,
    "white_level": 1023,
}

STANDARD_BAYER_PATTERNS = {"RGGB", "BGGR", "GRBG", "GBRG"}
QPD_CFA_PATTERN = "RGGB"
QPD_CFA_LAYOUT = "quad_bayer_2x2_blocks"
QPD_SIMULATOR_TYPE = "hwk_full_field"
QPD_SIMULATOR_VERSION = 1

SRGB_FROM_XYZ = np.asarray(
    [
        [3.2404542, -1.5371385, -0.4985314],
        [-0.9692660, 1.8760108, 0.0415560],
        [0.0556434, -0.2040259, 1.0572252],
    ],
    dtype=np.float32,
)
SRGB_TO_XYZ_D65 = np.linalg.inv(SRGB_FROM_XYZ).astype(np.float32)


def srgb_to_linear(srgb):
    srgb = np.clip(srgb, 0.0, 1.0)
    return np.where(srgb <= 0.04045, srgb / 12.92, ((srgb + 0.055) / 1.055) ** 2.4)


def linear_to_srgb(linear_rgb):
    linear_rgb = np.clip(linear_rgb, 0.0, 1.0)
    return np.where(linear_rgb <= 0.0031308, linear_rgb * 12.92, 1.055 * (linear_rgb ** (1.0 / 2.4)) - 0.055)


def load_srgb_image(path):
    bgr = cv2.imread(str(path), cv2.IMREAD_UNCHANGED)
    if bgr is None:
        raise FileNotFoundError(f"Cannot read image: {path}")
    if bgr.ndim == 2:
        bgr = np.repeat(bgr[..., None], 3, axis=2)
    if bgr.shape[2] == 4:
        bgr = bgr[:, :, :3]
    rgb = bgr[..., ::-1].astype(np.float32)
    max_value = 65535.0 if rgb.max() > 255.0 else 255.0
    return rgb / max_value


def cfa_pattern_from_rawpy(raw):
    if raw.raw_pattern.shape != (2, 2):
        raise ValueError(f"Unsupported RAW pattern shape {raw.raw_pattern.shape}; expected a standard 2x2 Bayer CFA.")
    color_desc = raw.color_desc.decode("ascii")
    return "".join(color_desc[int(raw.raw_pattern[y, x])] for y in range(2) for x in range(2)).upper()


def wb_gains_from_rawpy(raw):
    wb = np.asarray(raw.camera_whitebalance, dtype=np.float32)
    if wb.size < 4 or not np.all(np.isfinite(wb[:4])) or np.max(wb[:4]) <= 0:
        wb = np.asarray(raw.daylight_whitebalance, dtype=np.float32)

    g = np.mean([wb[1], wb[3]]) if wb.size >= 4 and wb[3] > 0 else wb[1]
    if g <= 0:
        return list(DEFAULT_ISP_PARAMS["wb_gains"])
    return [float(wb[0] / g), 1.0, float(wb[2] / g)]


def validate_ccm_matrix(ccm, name="ccm_srgb_from_cam"):
    ccm = np.asarray(ccm, dtype=np.float32)
    if ccm.shape != (3, 3):
        raise ValueError(f"{name} must be a 3x3 matrix, got shape {ccm.shape}")
    if not np.all(np.isfinite(ccm)):
        raise ValueError(f"{name} contains NaN or Inf values")
    det = float(np.linalg.det(ccm))
    if abs(det) < 1e-8:
        raise ValueError(f"{name} is singular or nearly singular; determinant={det:.3e}")
    return ccm


def metadata_ccm_from_rawpy(raw):
    info = {"method": "unavailable"}

    color_matrix = np.asarray(raw.color_matrix, dtype=np.float32)
    if color_matrix.size > 0:
        info["color_matrix_shape"] = list(color_matrix.shape)
    if color_matrix.ndim == 2 and color_matrix.shape[0] == 3 and color_matrix.shape[1] >= 3:
        try:
            ccm = validate_ccm_matrix(color_matrix[:, :3], name="raw.color_matrix[:, :3]")
        except ValueError as exc:
            info["color_matrix_error"] = str(exc)
        else:
            return ccm.tolist(), {
                "method": "raw.color_matrix[:, :3]",
                "source_shape": list(color_matrix.shape),
            }

    rgb_xyz_matrix = np.asarray(raw.rgb_xyz_matrix, dtype=np.float32)
    if rgb_xyz_matrix.size > 0:
        info["rgb_xyz_matrix_shape"] = list(rgb_xyz_matrix.shape)
    if rgb_xyz_matrix.ndim == 2 and rgb_xyz_matrix.shape[0] >= 3 and rgb_xyz_matrix.shape[1] == 3:
        cam_xyz = rgb_xyz_matrix[:3, :3]
        srgb_to_camera = cam_xyz @ SRGB_TO_XYZ_D65
        row_sums = np.sum(srgb_to_camera, axis=1, keepdims=True)
        if np.any(np.abs(row_sums) < 1e-8):
            info["rgb_xyz_matrix_error"] = "row sums are near zero"
        else:
            srgb_to_camera = srgb_to_camera / row_sums
            try:
                ccm = validate_ccm_matrix(np.linalg.pinv(srgb_to_camera), name="metadata CCM from raw.rgb_xyz_matrix")
            except ValueError as exc:
                info["rgb_xyz_matrix_error"] = str(exc)
            else:
                return ccm.tolist(), {
                    "method": "pinv(row_normalize(raw.rgb_xyz_matrix[:3, :3] @ srgb_to_xyz_d65))",
                    "source_shape": list(rgb_xyz_matrix.shape),
                }

    return None, info


def read_exif_iso(path):
    try:
        import exifread
    except ImportError:
        return None

    try:
        with open(path, "rb") as f:
            tags = exifread.process_file(f, details=False)
    except Exception:
        return None

    for key in ("EXIF ISOSpeedRatings", "EXIF PhotographicSensitivity", "Image ISOSpeedRatings"):
        value = tags.get(key)
        if value is None:
            continue
        try:
            raw_value = value.values[0] if hasattr(value, "values") else value
            return float(raw_value.num) / float(raw_value.den) if hasattr(raw_value, "num") else float(raw_value)
        except Exception:
            try:
                return float(str(value).split(",")[0].strip())
            except Exception:
                continue
    return None


def load_raw_with_isp_params(path, override_params=None):
    try:
        import rawpy
    except ImportError as exc:
        raise ImportError(
            "RAW/DNG input requires rawpy. Install rawpy, or use --input-kind srgb "
            "with --isp-json for the debug path."
        ) from exc

    with rawpy.imread(str(path)) as raw:
        raw_data = raw.raw_image_visible.astype(np.float32).copy()
        cfa_pattern = cfa_pattern_from_rawpy(raw)
        black_levels = [float(v) for v in raw.black_level_per_channel[:4]]
        cfa_black_levels = [
            float(raw.black_level_per_channel[int(raw.raw_pattern[y, x])])
            for y in range(2)
            for x in range(2)
        ]
        black_level = float(np.mean(black_levels)) if black_levels else 0.0
        white_level = float(raw.white_level) if raw.white_level is not None else float(np.max(raw_data))
        bit_depth = int(np.ceil(np.log2(max(white_level + 1.0, 2.0))))
        metadata_ccm, metadata_ccm_info = metadata_ccm_from_rawpy(raw)

        isp_params = dict(DEFAULT_ISP_PARAMS)
        isp_params.update(
            {
                "source_kind": "raw",
                "cfa_pattern": cfa_pattern,
                "black_level": black_level,
                "black_level_per_channel": black_levels,
                "cfa_black_level_2x2": cfa_black_levels,
                "white_level": white_level,
                "bit_depth": bit_depth,
                "iso": None,
                "iso_source": "unavailable",
                "wb_gains": wb_gains_from_rawpy(raw),
                "has_metadata_ccm": metadata_ccm is not None,
                "metadata_ccm_info": metadata_ccm_info,
                "ccm_srgb_from_cam": metadata_ccm
                if metadata_ccm is not None
                else DEFAULT_ISP_PARAMS["ccm_srgb_from_cam"],
            }
        )

        iso = getattr(raw, "iso_speed", None)
        if iso:
            isp_params["iso"] = float(iso)
            isp_params["iso_source"] = "rawpy.iso_speed"

    if isp_params.get("iso") is None:
        exif_iso = read_exif_iso(path)
        if exif_iso is not None:
            isp_params["iso"] = float(exif_iso)
            isp_params["iso_source"] = "exifread"

    if override_params:
        isp_params.update(override_params)
    return raw_data, isp_params


def load_rawpy_linear_srgb(path, crop_width=None, crop_height=None):
    try:
        import rawpy
    except ImportError as exc:
        raise ImportError("RAW/DNG linear sRGB rendering requires rawpy.") from exc

    with rawpy.imread(str(path)) as raw:
        srgb_u16 = raw.postprocess(
            use_camera_wb=True,
            output_color=rawpy.ColorSpace.sRGB,
            gamma=(1, 1),
            no_auto_bright=True,
            output_bps=16,
            user_flip=0,
        )
    linear_srgb = srgb_u16.astype(np.float32) / 65535.0
    return center_crop_quad(linear_srgb, crop_width, crop_height)


def normalize_raw_mosaic(raw_data, isp_params):
    raw_data = raw_data.astype(np.float32)
    pattern = validate_cfa_pattern(isp_params["cfa_pattern"])
    black_default = float(isp_params["black_level"])
    white_level = float(isp_params["white_level"])
    black_levels = isp_params.get("cfa_black_level_2x2") or isp_params.get("black_level_per_channel")

    if black_levels and len(black_levels) >= 4:
        black_map = np.empty(raw_data.shape, dtype=np.float32)
        for y in range(2):
            for x in range(2):
                black_map[y::2, x::2] = float(black_levels[y * 2 + x])
    else:
        black_map = black_default

    denom = np.maximum(white_level - black_map, 1.0)
    return np.clip((raw_data - black_map) / denom, 0.0, 1.0).astype(np.float32)


def demosaic_raw_to_camera_rgb(raw_linear, cfa_pattern):
    pattern = validate_cfa_pattern(cfa_pattern)
    codes = {
        "RGGB": cv2.COLOR_BayerBG2RGB,
        "BGGR": cv2.COLOR_BayerRG2RGB,
        "GRBG": cv2.COLOR_BayerGB2RGB,
        "GBRG": cv2.COLOR_BayerGR2RGB,
    }
    if pattern not in codes:
        raise ValueError(f"Unsupported CFA pattern {pattern}; expected one of {sorted(codes)}.")
    raw_u16 = np.rint(np.clip(raw_linear, 0.0, 1.0) * 65535.0).astype(np.uint16)
    return (cv2.cvtColor(raw_u16, codes[pattern]).astype(np.float32) / 65535.0).clip(0.0, 1.0)


def apply_forward_isp_linear(camera_rgb, isp_params):
    wb_gains = np.asarray(isp_params["wb_gains"], dtype=np.float32).reshape(1, 1, 3)
    balanced = camera_rgb * wb_gains
    ccm = validate_ccm_matrix(isp_params["ccm_srgb_from_cam"])
    return (balanced @ ccm.T).astype(np.float32)


def inverse_linear_isp_to_camera_rgb(linear_srgb, isp_params):
    ccm = validate_ccm_matrix(isp_params["ccm_srgb_from_cam"])
    cam_from_srgb = np.linalg.inv(ccm)
    balanced = linear_srgb @ cam_from_srgb.T
    wb_gains = np.asarray(isp_params["wb_gains"], dtype=np.float32).reshape(1, 1, 3)
    return (balanced / np.maximum(wb_gains, 1e-6)).astype(np.float32)


def raw_to_clean_energy_field(raw_data, isp_params):
    raw_linear = normalize_raw_mosaic(raw_data, isp_params)
    return demosaic_raw_to_camera_rgb(raw_linear, isp_params["cfa_pattern"])


def fit_ccm_from_linear_srgb(clean_energy_field, reference_linear_srgb, isp_params, max_samples=300000):
    clean_shape = clean_energy_field.shape
    reference_shape = reference_linear_srgb.shape
    if clean_shape[:2] != reference_shape[:2]:
        fit_w = min(clean_shape[1], reference_shape[1])
        fit_h = min(clean_shape[0], reference_shape[0])
        clean_energy_field = center_crop_quad(clean_energy_field, fit_w, fit_h)
        reference_linear_srgb = center_crop_quad(reference_linear_srgb, fit_w, fit_h)

    wb_gains = np.asarray(isp_params["wb_gains"], dtype=np.float32).reshape(1, 1, 3)
    balanced = clean_energy_field * wb_gains
    x = balanced.reshape(-1, 3)
    y = reference_linear_srgb.reshape(-1, 3)

    mask = (
        np.all(np.isfinite(x), axis=1)
        & np.all(np.isfinite(y), axis=1)
        & (np.max(x, axis=1) > 1e-4)
        & (np.max(x, axis=1) < 0.95)
        & (np.max(y, axis=1) > 1e-4)
        & (np.max(y, axis=1) < 0.95)
    )
    idx = np.flatnonzero(mask)
    if idx.size == 0:
        raise ValueError("No valid pixels for CCM fitting")
    if idx.size > max_samples:
        rng = np.random.default_rng(12345)
        idx = rng.choice(idx, size=max_samples, replace=False)

    a, *_ = np.linalg.lstsq(x[idx], y[idx], rcond=None)
    ccm = validate_ccm_matrix(a.T.astype(np.float32), name="fitted ccm_srgb_from_cam")
    pred = np.clip(balanced @ ccm.T, 0.0, 1.0)
    mae = np.mean(np.abs(pred - reference_linear_srgb), axis=(0, 1))
    return ccm, {
        "fit_pixels": int(idx.size),
        "clean_shape": list(clean_shape),
        "reference_shape": list(reference_shape),
        "fit_shape": list(clean_energy_field.shape),
        "determinant": float(np.linalg.det(ccm)),
        "reference": "rawpy linear sRGB, use_camera_wb=True, gamma=(1,1), no_auto_bright=True",
        "mae_rgb": [float(v) for v in mae],
        "mae_mean": float(np.mean(mae)),
    }


def inverse_isp_to_camera_rgb(srgb, isp_params):
    linear_srgb = srgb_to_linear(srgb)
    ccm = validate_ccm_matrix(isp_params["ccm_srgb_from_cam"])
    cam_from_srgb = np.linalg.inv(ccm)
    cam_rgb = linear_srgb @ cam_from_srgb.T

    wb_gains = np.asarray(isp_params["wb_gains"], dtype=np.float32).reshape(1, 1, 3)
    cam_rgb = cam_rgb / np.maximum(wb_gains, 1e-6)
    return np.clip(cam_rgb, 0.0, 1.0).astype(np.float32)


def validate_cfa_pattern(cfa_pattern):
    pattern = str(cfa_pattern).upper()
    if pattern not in STANDARD_BAYER_PATTERNS:
        raise ValueError(f"cfa_pattern must be one of {sorted(STANDARD_BAYER_PATTERNS)}, got {cfa_pattern!r}")
    return pattern


def center_crop_quad(array, width, height):
    if width is None and height is None:
        return array
    if width is None or height is None:
        raise ValueError("--crop must be WIDTHxHEIGHT, for example 3000x2000")

    h, w = array.shape[:2]
    crop_w = min(width, w)
    crop_h = min(height, h)
    crop_w -= crop_w % 4
    crop_h -= crop_h % 4
    if crop_w < 4 or crop_h < 4:
        raise ValueError("Quad RGGB crop dimensions must both be at least 4 pixels")
    x0 = ((w - crop_w) // 2) // 4 * 4
    y0 = ((h - crop_h) // 2) // 4 * 4
    return array[y0:y0 + crop_h, x0:x0 + crop_w, ...]


def parse_crop(crop):
    if crop is None:
        return None, None
    parts = crop.lower().replace("*", "x").split("x")
    if len(parts) != 2:
        raise ValueError("--crop must be WIDTHxHEIGHT, for example 3000x2000")
    width, height = int(parts[0]), int(parts[1])
    if width < 4 or height < 4:
        raise ValueError("--crop dimensions must both be at least 4 pixels")
    return width, height


def load_noise_row(noise_table_path, iso, rng):
    noise_table_path = Path(noise_table_path)
    if noise_table_path.suffix.lower() == ".csv":
        with open(noise_table_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            rows = list(reader)
        for row in rows:
            for key, value in list(row.items()):
                try:
                    row[key] = float(value)
                except (TypeError, ValueError):
                    pass
    else:
        try:
            import openpyxl
        except ImportError as exc:
            fallback = noise_table_path.with_suffix(".csv")
            if fallback.exists():
                return load_noise_row(fallback, iso, rng)
            raise ImportError(
                "Reading .xlsx noise tables requires openpyxl. "
                "Install openpyxl or provide a .csv noise table."
            ) from exc

        wb = openpyxl.load_workbook(noise_table_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [str(ws.cell(1, c).value).strip() for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            item = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
            if item.get("ISO") is not None:
                rows.append(item)
    if not rows:
        raise ValueError(f"No ISO rows found in {noise_table_path}")

    required_columns = {"ISO", "k-10bit", "b-10bit"}
    missing_columns = required_columns.difference(rows[0].keys())
    if missing_columns:
        raise ValueError(f"Noise table {noise_table_path} is missing columns: {sorted(missing_columns)}")

    normalized_rows = []
    for row_index, row in enumerate(rows, start=2):
        out = dict(row)
        try:
            out["ISO"] = float(row["ISO"])
            out["k-10bit"] = float(row["k-10bit"])
            out["b-10bit"] = float(row["b-10bit"])
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Noise table row {row_index} must have numeric ISO, k-10bit, and b-10bit values"
            ) from exc
        normalized_rows.append(out)

    rows = sorted(normalized_rows, key=lambda row: row["ISO"])
    available_iso = [float(row["ISO"]) for row in rows]

    def random_row(reason):
        idx = int(rng.integers(0, len(rows)))
        out = dict(rows[idx])
        out["requested_iso"] = None if iso is None else float(iso)
        out["noise_iso_selection"] = reason
        out["available_iso_min"] = min(available_iso)
        out["available_iso_max"] = max(available_iso)
        return out

    if iso is None:
        return random_row("missing_iso_random")

    iso = float(iso)
    exact = [row for row in rows if abs(float(row["ISO"]) - iso) < 1e-6]
    if exact:
        out = dict(exact[0])
        out["requested_iso"] = iso
        out["noise_iso_selection"] = "exact_iso_match"
        return out

    if iso < min(available_iso) or iso > max(available_iso):
        return random_row("out_of_range_random")

    return random_row("not_in_table_random")


def add_poisson_gaussian_noise(raw_linear, noise_row, black_level, white_level, bit_depth, rng):
    if bit_depth != 10:
        raise ValueError("The current noise table is fixed to the 10bit DN domain.")
    signal_dn = np.clip(raw_linear, 0.0, 1.0) * (white_level - black_level)

    # Assumption: the table stores 10-bit variance model parameters, variance = k * signal + b.
    k = float(noise_row["k-10bit"])
    b = float(noise_row["b-10bit"])
    variance = np.maximum(k * signal_dn + b, 0.0)
    noisy_signal_dn = signal_dn + rng.normal(0.0, np.sqrt(variance), size=signal_dn.shape)
    return np.clip(noisy_signal_dn / max(white_level - black_level, 1), 0.0, 1.0).astype(np.float32)


def quantize_raw(raw_linear, black_level, white_level, bit_depth):
    full_scale = (1 << bit_depth) - 1
    raw_dn = black_level + np.clip(raw_linear, 0.0, 1.0) * (white_level - black_level)
    return np.clip(np.rint(raw_dn), 0, full_scale).astype(np.uint16)


def save_preview(raw_quantized, black_level, white_level, output_path):
    preview = (raw_quantized.astype(np.float32) - black_level) / max(white_level - black_level, 1)
    preview = np.clip(preview, 0.0, 1.0)
    cv2.imwrite(str(output_path), np.rint(preview * 65535.0).astype(np.uint16))


def save_srgb_image(srgb, output_path):
    srgb = np.nan_to_num(srgb, nan=0.0, posinf=1.0, neginf=0.0)
    rgb_u8 = np.rint(np.clip(srgb, 0.0, 1.0) * 255.0).astype(np.uint8)
    cv2.imwrite(str(output_path), rgb_u8[..., ::-1])


def save_linear_rgb_preview(linear_rgb, output_path):
    preview = np.clip(linear_rgb, 0.0, 1.0)
    save_srgb_image(linear_to_srgb(preview), output_path)


def load_json_or_default(path, default):
    if path is None:
        return dict(default)
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    merged = dict(default)
    merged.update(data)
    return merged


def sha256_file(path):
    if path is None:
        return None
    digest = hashlib.sha256()
    with open(path, "rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main():
    parser = argparse.ArgumentParser(description="RAW/sRGB ISP round-trip to full-field HWK QPD raw simulator")
    parser.add_argument("--input", required=True, help="Input RAW/DNG or sRGB image path")
    parser.add_argument(
        "--input-kind",
        choices=("auto", "raw", "srgb"),
        default="auto",
        help="raw reads Bayer data and ISP params from RAW metadata; srgb uses --isp-json/default params",
    )
    parser.add_argument("--output-dir", default="outputs", help="Output directory")
    parser.add_argument("--isp-json", help="Optional ISP parameter JSON overrides; sRGB mode uses defaults plus this file")
    parser.add_argument("--hwk-dir", help="HWK field_data directory or statistics root")
    parser.add_argument("--hwk-config", help="Optional HWK/RDM simulator config JSON")
    parser.add_argument("--hwk-distance", help="Select one calibrated object distance")
    parser.add_argument("--hwk-aperture", help="Select one calibrated aperture")
    parser.add_argument("--no-hwk-cache", action="store_true", help="Disable adjacent .csv.npz HWK caches")
    parser.add_argument("--noise-table", default="noise_table.csv", help="Noise table .xlsx or .csv path")
    parser.add_argument("--iso", type=float, help="Override ISO used for noise table lookup")
    parser.add_argument("--seed", type=int, default=0)
    parser.add_argument("--crop", default="3000x2000", help="4-pixel-aligned center crop before ISP/QPD simulation, default 3000x2000")
    parser.add_argument(
        "--ccm-source",
        choices=("auto", "rawpy-fit", "metadata", "identity"),
        default="metadata",
        help="metadata uses raw.color_matrix or derives it from raw.rgb_xyz_matrix; rawpy-fit fits to rawpy linear sRGB.",
    )
    parser.add_argument("--skip-qpd-sim", action="store_true", help="Skip HWK/RDM while retaining Quad RGGB sampling")
    parser.add_argument("--skip-noise", action="store_true")
    args = parser.parse_args()

    if not args.skip_qpd_sim and args.hwk_dir is None:
        parser.error("--hwk-dir is required unless --skip-qpd-sim is used")

    seed_sequence = np.random.SeedSequence(args.seed)
    simulator_sequence, noise_sequence = seed_sequence.spawn(2)
    simulator_seed = int(simulator_sequence.generate_state(1, dtype=np.uint32)[0])
    noise_seed = int(noise_sequence.generate_state(1, dtype=np.uint32)[0])
    noise_rng = np.random.default_rng(noise_sequence)

    input_path = Path(args.input)
    raw_suffixes = {
        ".dng",
        ".arw",
        ".cr2",
        ".cr3",
        ".nef",
        ".nrw",
        ".raf",
        ".raw",
        ".rw2",
        ".orf",
        ".pef",
        ".srw",
    }
    input_kind = args.input_kind
    if input_kind == "auto":
        input_kind = "raw" if input_path.suffix.lower() in raw_suffixes else "srgb"

    override_isp_params = None
    if args.isp_json is not None:
        with open(args.isp_json, "r", encoding="utf-8") as f:
            override_isp_params = json.load(f)

    crop_width, crop_height = parse_crop(args.crop)
    linear_srgb = None
    roundtrip_error = None
    ccm_fit_diagnostics = None
    if input_kind == "raw":
        raw_data, isp_params = load_raw_with_isp_params(input_path, override_isp_params)
        raw_data = center_crop_quad(raw_data, crop_width, crop_height)
        clean_energy_field = raw_to_clean_energy_field(raw_data, isp_params)
        requested_ccm_source = args.ccm_source
        has_metadata_ccm = bool(isp_params.get("has_metadata_ccm"))
        if requested_ccm_source == "auto":
            requested_ccm_source = "metadata" if has_metadata_ccm else "identity"

        if requested_ccm_source == "rawpy-fit":
            reference_linear_srgb = load_rawpy_linear_srgb(input_path, crop_width, crop_height)
            fitted_ccm, ccm_fit_diagnostics = fit_ccm_from_linear_srgb(
                clean_energy_field, reference_linear_srgb, isp_params
            )
            isp_params["ccm_srgb_from_cam"] = fitted_ccm.tolist()
            isp_params["ccm_source"] = "rawpy-fit"
        elif requested_ccm_source == "identity":
            isp_params["ccm_srgb_from_cam"] = np.eye(3, dtype=np.float32).tolist()
            isp_params["ccm_source"] = "identity"
        else:
            override_has_ccm = override_isp_params is not None and "ccm_srgb_from_cam" in override_isp_params
            if override_has_ccm:
                isp_params["ccm_source"] = "provided_isp_json_ccm"
            elif has_metadata_ccm:
                isp_params["ccm_source"] = "metadata"
            else:
                raise ValueError(
                    "--ccm-source metadata requires raw.color_matrix, raw.rgb_xyz_matrix, "
                    "or a 3x3 --isp-json ccm_srgb_from_cam override. Use --ccm-source identity to bypass CCM."
                )
        isp_params["ccm_srgb_from_cam"] = validate_ccm_matrix(isp_params["ccm_srgb_from_cam"]).tolist()
        isp_params["ccm_source_requested"] = args.ccm_source

        linear_srgb = apply_forward_isp_linear(clean_energy_field, isp_params)
        srgb = linear_to_srgb(linear_srgb).astype(np.float32)
        roundtrip_camera_rgb = inverse_linear_isp_to_camera_rgb(linear_srgb, isp_params)
        roundtrip_diff = roundtrip_camera_rgb - clean_energy_field
        roundtrip_error = {
            "max_abs": float(np.max(np.abs(roundtrip_diff))),
            "mean_abs": float(np.mean(np.abs(roundtrip_diff))),
        }
    else:
        isp_params = load_json_or_default(args.isp_json, DEFAULT_ISP_PARAMS)
        isp_params["source_kind"] = "srgb"
        isp_params["ccm_source"] = "provided_srgb_params"
        isp_params["ccm_source_requested"] = args.ccm_source
        isp_params["ccm_srgb_from_cam"] = validate_ccm_matrix(isp_params["ccm_srgb_from_cam"]).tolist()
        srgb = load_srgb_image(input_path)
        srgb = center_crop_quad(srgb, crop_width, crop_height)
        clean_energy_field = inverse_isp_to_camera_rgb(srgb, isp_params)
        linear_srgb = apply_forward_isp_linear(clean_energy_field, isp_params)
        roundtrip_error = {
            "max_abs": None,
            "mean_abs": None,
            "note": "sRGB input is not guaranteed reversible because it may already be gamma encoded, clipped, and quantized.",
        }

    hwk_config = load_json_or_default(args.hwk_config, {})
    if args.iso is not None:
        isp_params["iso"] = float(args.iso)
        isp_params["iso_source"] = "command_line_override"

    input_bit_depth = int(isp_params["bit_depth"])
    input_black_level = float(isp_params["black_level"])
    input_white_level = float(isp_params["white_level"])
    output_bit_depth = int(QPD_OUTPUT_LEVELS["bit_depth"])
    output_black_level = float(QPD_OUTPUT_LEVELS["black_level"])
    output_white_level = float(QPD_OUTPUT_LEVELS["white_level"])
    iso = None if isp_params.get("iso") is None else float(isp_params["iso"])

    qpd_camera_rgb = np.clip(clean_energy_field, 0.0, 1.0)
    cfa_planes = quad_rggb_sample_to_planes(qpd_camera_rgb)
    if args.skip_qpd_sim:
        simulated_planes = cfa_planes
        qpd_simulation_meta = {
            "applied": False,
            "input_domain": "linear_camera_cfa_planes",
            "cfa_order": ["R", "Gr", "Gb", "B"],
        }
        simulator_config = None
    else:
        hwk_bank = HwkBank(args.hwk_dir, use_cache=not args.no_hwk_cache)
        simulator = QpdHwkSimulator(
            hwk_bank,
            config=hwk_config,
            seed=simulator_seed,
        )
        simulated_planes, qpd_simulation_meta = simulator(
            cfa_planes,
            distance=args.hwk_distance,
            aperture=args.hwk_aperture,
            return_meta=True,
        )
        qpd_simulation_meta["applied"] = True
        simulator_config = simulator.config

    raw_linear = quad_rggb_planes_to_mosaic(simulated_planes)

    noise_row = None
    if not args.skip_noise:
        noise_row = load_noise_row(args.noise_table, iso, noise_rng)
        raw_linear = add_poisson_gaussian_noise(
            raw_linear,
            noise_row,
            output_black_level,
            output_white_level,
            output_bit_depth,
            noise_rng,
        )

    raw_quantized = quantize_raw(raw_linear, output_black_level, output_white_level, output_bit_depth)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    np.save(output_dir / "clean_energy_field.npy", clean_energy_field.astype(np.float32))
    np.save(output_dir / "isp_linear_srgb.npy", linear_srgb.astype(np.float32))
    np.save(output_dir / "qpd_raw.npy", raw_quantized)
    save_srgb_image(srgb, output_dir / "isp_srgb.png")
    save_linear_rgb_preview(clean_energy_field, output_dir / "clean_energy_preview.png")
    save_preview(raw_quantized, output_black_level, output_white_level, output_dir / "qpd_raw_preview.png")

    metadata = {
        "input": str(args.input),
        "input_kind": input_kind,
        "crop": args.crop,
        "isp_srgb_source": "linear_to_srgb(clip(isp_linear_srgb)) reversible ISP preview",
        "isp_srgb": str(output_dir / "isp_srgb.png"),
        "clean_energy_field": str(output_dir / "clean_energy_field.npy"),
        "isp_linear_srgb": str(output_dir / "isp_linear_srgb.npy"),
        "clean_energy_domain": "linear camera RGB after black-level subtraction, white-level normalization, and demosaic",
        "reversible_isp_contract": "clean_energy_field -> AWB -> CCM -> isp_linear_srgb; inverse uses inverse(CCM) and inverse AWB on float tensors before gamma/clip/quantization",
        "reversible_isp_roundtrip_error": roundtrip_error,
        "ccm_fit_diagnostics": ccm_fit_diagnostics,
        "shape": list(raw_quantized.shape),
        "qpd_raw_shape": list(raw_quantized.shape),
        "clean_energy_shape": list(clean_energy_field.shape),
        "isp_linear_srgb_shape": list(linear_srgb.shape),
        "dtype": str(raw_quantized.dtype),
        "input_raw_levels": {
            "bit_depth": input_bit_depth,
            "black_level": input_black_level,
            "white_level": input_white_level,
        },
        "qpd_output_levels": {
            "bit_depth": output_bit_depth,
            "black_level": output_black_level,
            "white_level": output_white_level,
        },
        "qpd_readout_mode": "same",
        "qpd_readout_scale": 1,
        "qpd_cfa_pattern": QPD_CFA_PATTERN,
        "qpd_cfa_layout": QPD_CFA_LAYOUT,
        "qpd_simulator_type": "disabled" if args.skip_qpd_sim else QPD_SIMULATOR_TYPE,
        "qpd_simulator_version": QPD_SIMULATOR_VERSION,
        "qpd_simulation_request": {
            "hwk_dir": None if args.hwk_dir is None else str(Path(args.hwk_dir).resolve()),
            "hwk_config": None if args.hwk_config is None else str(Path(args.hwk_config).resolve()),
            "hwk_config_sha256": sha256_file(args.hwk_config),
            "distance": args.hwk_distance,
            "aperture": args.hwk_aperture,
            "cache_enabled": not args.no_hwk_cache,
        },
        "qpd_simulator_config": simulator_config,
        "qpd_simulation": qpd_simulation_meta,
        "simulation_seed": simulator_seed,
        "noise_seed": noise_seed,
        "isp_params": isp_params,
        "noise_row": noise_row,
    }
    with open(output_dir / "metadata.json", "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)

    print(f"saved {output_dir / 'qpd_raw.npy'}")
    print(f"saved {output_dir / 'qpd_raw_preview.png'}")
    print(f"saved {output_dir / 'metadata.json'}")


if __name__ == "__main__":
    main()
